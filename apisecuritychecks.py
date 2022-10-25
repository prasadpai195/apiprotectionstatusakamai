import requests
import json
from akamai.edgegrid import EdgeGridAuth, EdgeRc
import os
from urllib.parse  import urljoin
from openpyxl import load_workbook
import openpyxl, pprint
import sys

#Enter API creds below
edgerc_file = os.path.join(os.path.expanduser("~"), '.edgerc')
edgerc = EdgeRc(edgerc_file)
section="[ENTER YOUR EDGERC SECTION HERE]"
base_url = edgerc.get(section,'host')
baseurl=str('https://')+str(base_url)
client_token=edgerc.get(section,'client_token')
client_secret=edgerc.get(section,'client_secret')
access_token=edgerc.get(section,'access_token')
s = requests.Session()
s.auth = EdgeGridAuth(
client_token=client_token,
client_secret=client_secret,
access_token=access_token
)
if __name__ == '__main__':
	wb = load_workbook('input.xlsx')
	sheet = wb['Sheet1']
	sheet2=wb['SheetX']
	print("Number of entries is:" ,sheet.max_row-1, "hostnames")
	i=2
	for row in range(2, sheet.max_row+1):
		hostname=sheet['A' + str(row)].value
		print("*************************************************************************************************************************************************************")
		print("Running for "+ hostname  +"")
		skey="[CLEANUP THE SKEY REFERENCES IN THE SCRIPT IF YOU ARE RUNNING THIS SCRIPT FOR YOUR ACCOUNT ONLY]"
		# PULL CONFIGURATIONS
		config=s.get(baseurl + ("/appsec/v1/configs?accountSwitchKey="+skey+"") , headers = {'PAPI-Use-Prefixes': 'true'})
		configs = json.loads(config.text)
		configlist = (configs['configurations'])
		for config in configlist:
			configid= str(config['id'])
			configname=str(config['name'])
			prodversion=str(config['productionVersion'])
			prodhostnames=config['productionHostnames']
			#print(prodhostnames)
			for hostnames in prodhostnames:
				if hostname == str(hostnames):
					print("hostname " +hostname+ "found in "+configname+"")
					mt=s.get(baseurl + ("/appsec/v1/configs/"+configid+"/versions/"+prodversion+"/hostname-coverage/match-targets?accountSwitchKey="+skey+"&hostname="+hostname+""), headers = {'PAPI-Use-Prefixes': 'true'})
					print(mt.json())
					mtjson=json.loads(mt.text)
					targets=mtjson['matchTargets']
					print(targets['websiteTargets'])
					if len(targets['websiteTargets']) != 0:
						for webtargets in targets['websiteTargets']:
							try:
								wafwcontrols=str(webtargets['effectiveSecurityControls']['applyApplicationLayerControls'])
							except:
								wafwconstraints="false"
							try:
								botwcontrols=str(webtargets['effectiveSecurityControls']['applyBotmanControls'])
							except:
								botwconstraints="false"
							try:
								nlwcontrols=str(webtargets['effectiveSecurityControls']['applyNetworkLayerControls'])
							except:
								nlwconstraints="false"
							try:
								ratewcontrols=str(webtargets['effectiveSecurityControls']['applyRateControls'])
							except:
								ratewconstraints="false"
							try:
								crwcontrols=str(webtargets['effectiveSecurityControls']['applyReputationControls'])
							except:
								crwconstraints="false"
							try:
								slowpostwcontrols=str(webtargets['effectiveSecurityControls']['applySlowPostControls'])
							except:
								slowpostwconstraints="false"
							try:
								apiwconstraints=str(webtargets['effectiveSecurityControls']['applyApiConstraints'])
							except:
								apiwconstraints="false"
							try:
								policyw=str(webtargets['securityPolicy']['policyId'])
							except:
								policyw="false"
							#status=str("Policy="+policyw+", WAF="+wafwcontrols+", BOT="+botwcontrols+", NL="+nlwcontrols+", DOS="+ratewcontrols+", CR="+crwcontrols+", SLP="+slowpostwcontrols+", APIcons="+apiwconstraints+"")
							sheet2['A' + str(i)].value=str(hostname)
							sheet2['B' + str(i)].value=str(policyw)
							sheet2['C' + str(i)].value=str(wafwcontrols)
							sheet2['D' + str(i)].value=str(botwcontrols)
							sheet2['E' + str(i)].value=str(nlwcontrols)
							sheet2['F' + str(i)].value=str(ratewcontrols)
							sheet2['G' + str(i)].value=str(crwcontrols)
							sheet2['H' + str(i)].value=str(slowpostwcontrols)
							sheet2['I' + str(i)].value=str(apiwconstraints)
							wb.save("input.xlsx")
							i+=1;
					else:
							sheet2['A' + str(i)].value=str(hostname)
							sheet2['B' + str(i)].value="false"
							sheet2['C' + str(i)].value="false"
							sheet2['D' + str(i)].value="false"
							sheet2['E' + str(i)].value="false"
							sheet2['F' + str(i)].value="false"
							sheet2['G' + str(i)].value="false"
							sheet2['H' + str(i)].value="false"
							sheet2['I' + str(i)].value="false"
							wb.save("input.xlsx")
							i+=1;
						
					if len(targets['apiTargets'])!= "0":
						for apitargets in targets['apiTargets']:
							try:
								wafacontrols=str(apitargets['effectiveSecurityControls']['applyApplicationLayerControls'])
							except:
								wafacontrols="false"
							try:
								botacontrols=str(apitargets['effectiveSecurityControls']['applyBotmanControls'])
							except:
								botacontrols="false"
							try:
								nlacontrols=str(apitargets['effectiveSecurityControls']['applyNetworkLayerControls'])
							except:
								nlacontrols="false"
							try:
								rateacontrols=str(apitargets['effectiveSecurityControls']['applyRateControls'])
							except:
								rateacontrols="false"
							try:
								cracontrols=str(apitargets['effectiveSecurityControls']['applyReputationControls'])
							except:
								cracontrols="false"
							try:
								slowpostacontrols=str(apitargets['effectiveSecurityControls']['applySlowPostControls'])
							except:
								slowpostacontrols="false"
							try:
								apiaconstraints=str(apitargets['effectiveSecurityControls']['applyApiConstraints'])
							except:
								apiaconstraints="false"
							try:
								policya=str(apitargets['securityPolicy']['policyId'])
							except:
								policya="false"
							#status=str("Policy="+policya+", WAF="+wafacontrols+", BOT="+botacontrols+", NL="+nlacontrols+", DOS="+rateacontrols+", CR="+cracontrols+", SLP="+slowpostacontrols+", APIcons="+apiaconstraints+"")
							sheet2['A' + str(i)].value=str(hostname)
							sheet2['B' + str(i)].value=str(policya)
							sheet2['C' + str(i)].value=str(wafacontrols)
							sheet2['D' + str(i)].value=str(botacontrols)
							sheet2['E' + str(i)].value=str(nlacontrols)
							sheet2['F' + str(i)].value=str(rateacontrols)
							sheet2['G' + str(i)].value=str(cracontrols)
							sheet2['H' + str(i)].value=str(slowpostacontrols)
							sheet2['I' + str(i)].value=str(apiaconstraints)
							wb.save("input.xlsx")
							i+=1;				
					else:
							sheet2['A' + str(i)].value=str(hostname)
							sheet2['B' + str(i)].value="false"
							sheet2['C' + str(i)].value="false"
							sheet2['D' + str(i)].value="false"
							sheet2['E' + str(i)].value="false"
							sheet2['F' + str(i)].value="false"
							sheet2['G' + str(i)].value="false"
							sheet2['H' + str(i)].value="false"
							sheet2['I' + str(i)].value="false"
							wb.save("input.xlsx")
							i+=1;






		